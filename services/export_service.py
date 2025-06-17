import json
import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from services.search_service import get_user_searches, get_search_results  # убедись, что импортировал get_search_results

def export_user_searches(user_id: int, filename: str, format: str = "json", search_id: str = None) -> str:
    path = Path("storage")
    path.mkdir(exist_ok=True)
    file_path = path / f"{filename}.{format}"

    if format == "json":
        searches = get_user_searches(str(user_id))
        with open(file_path, "w", encoding="utf-8") as f:
            json.dump(searches, f, indent=4, ensure_ascii=False)

    elif format == "csv":
        searches = get_user_searches(str(user_id))
        with open(file_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["platform", "brand", "region", "min_price", "max_price", "notifications"])
            for platform, search_list in searches.items():
                for s in search_list:
                    p = s["params"]
                    writer.writerow([platform, p["brand"], p["region"], p["min_price"], p["max_price"], s["notifications"]])

    elif format == "xlsx":
        if not search_id:
            raise ValueError("search_id is required for Excel export")

        results = get_search_results(user_id, search_id)
        wb = Workbook()
        ws = wb.active
        ws.title = "Search Results"

        # Заголовки
        headers = ["Цена", "Название", "Пробег", "Год", "Дата публикации", "Ссылка"]
        ws.append(headers)

        # Стили для заголовков
        header_font = Font(bold=True)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col_num, column_title in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.alignment = center_align

        # Данные
        for ad in results:
            ws.append([
                ad.get("price", "—"),
                ad.get("title", "—"),
                ad.get("mileage", "—"),
                ad.get("year", "—"),
                ad.get("date", "—"),
                ad.get("url", "—")
            ])

        # Автоширина колонок
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
            col_letter = get_column_letter(column_cells[0].column)
            ws.column_dimensions[col_letter].width = length + 4

        # Автофильтр
        ws.auto_filter.ref = ws.dimensions

        # Заморозка заголовка
        ws.freeze_panes = "A2"

        wb.save(file_path)

    else:
        raise ValueError(f"Unsupported export format: {format}")

    return str(file_path)
